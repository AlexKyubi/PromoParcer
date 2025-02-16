import requests
import pathlib
import openpyxl
from io import BytesIO
from openpyxl.styles import Font
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from tqdm import tqdm
import subprocess


class PromoParser:
    """Класс для парсинга акций с сайтов Sulpak, Mechta и Technodom."""



    def __init__(self, output_file="promotions.xlsx"):
        """
        Инициализирует объект парсера.
        :param output_file: Путь к файлу, в который сохраняются акции.
        """
        self.SULPAK_URL = "https://www.sulpak.kz/Stocks"
        self.MECHTA_URL = "https://www.mechta.kz/useful/shares/"
        self.TECHNODOM_URL = "https://www.technodom.kz/cms/promo"
        self.file_path = pathlib.Path(output_file)

        # Настройки Selenium (без открытия браузера)
        self.chrome_options = Options()
        self.chrome_options.add_argument("--headless")
        self.chrome_options.add_argument("--disable-gpu")
        self.chrome_options.add_argument("--no-sandbox")
        self.chrome_options.add_argument("--disable-dev-shm-usage")
        self.chrome_options.add_argument("--log-level=3")  # Отключаем лишние логи
        self.chrome_options.add_argument("--silent")  # Делаем браузер тихим
        self.chrome_options.add_argument("--disable-logging")  # Запрещаем выводить логи
        self.chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # Маскируем Selenium
        self.chrome_options.add_argument("--disable-webgl")
        self.chrome_options.add_argument("--disable-gpu-compositing")
        self.chrome_options.add_argument("--disable-software-rasterizer")
        self.chrome_options.add_argument("--disable-3d-apis")
        self.chrome_options.add_argument("--remote-debugging-port=0")  # Убираем DevTools logging
        self.chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36")
        self.service = Service(ChromeDriverManager().install(), log_output=subprocess.DEVNULL)

     
    

    def fetch_html(self,url):
        """Получает HTML-код страницы (только для Sulpak)."""
        try:
            response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
            response.raise_for_status()
            return response.text
        except requests.exceptions.Timeout:
            print(f"❌ Тайм-аут при подключении к {url}")
        except requests.exceptions.RequestException as e:
            print(f"❌ Ошибка соединения: {e}")
        return None

    def parse_sulpak(self):
        """Парсит акции с Sulpak.kz."""
        print(f"[🔍 Парсинг Sulpak] {self.SULPAK_URL}")

        html = self.fetch_html(self.SULPAK_URL)
        if not html:
            return []

        soup = BeautifulSoup(html, "html.parser")
        promotions = []

        promo_blocks = soup.find_all("div", class_="actions__item-name")

        for promo in tqdm(promo_blocks, desc="Обрабатываем акции Sulpak"):
            title = promo.get_text(strip=True)  # Название акции
            description_tag = promo.find_next("div", class_="actions__item-text")  # Описание
            period_tag = promo.find_next("div", class_="actions__item-period")  # Период

            if title:
                promotions.append([
                    title,
                    description_tag.get_text(strip=True) if description_tag else "Нет описания",
                    period_tag.get_text(strip=True) if period_tag else "Не указано"
                ])

        return promotions

    def parse_mechta(self):
        """Парсит акции с Mechta.kz (через Selenium) с переключением страниц по параметру ?page=."""
        print(f"[🔍 Парсинг Mechta] {self.MECHTA_URL}")

        try:
            driver = webdriver.Chrome(service=self.service, options=self.chrome_options)
            all_promotions = []
            page = 1  # Начинаем с первой страницы
            MAX_PAGES = 5  # Ограничение, чтобы не зациклиться
            last_page_source = ""  # Для проверки повторяющихся страниц

            while page <= MAX_PAGES:  # Не даем зациклиться бесконечно
                url = f"{self.MECHTA_URL}?page={page}"
                print(f"🔄 Загружаем страницу {page}: {url}")
                driver.get(url)

                # Ждём загрузку хотя бы одного блока с акцией
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "column"))
                    )
                except:
                    print("🚫 Больше страниц нет, останавливаемся.")
                    break  # Если элементы не появились, значит, страницы закончились

                html = driver.page_source

                # Проверяем, не повторяется ли страница (защита от бесконечного цикла)
                if html == last_page_source:
                    print("🚫 Повторяющаяся страница! Останавливаем парсинг.")
                    break
                last_page_source = html  # Запоминаем текущую страницу

                soup = BeautifulSoup(html, "html.parser")

                # 🔹 Ищем акции на текущей странице
                promo_blocks = soup.find_all("div", class_="column")

                if not promo_blocks:
                    print("🚫 Акции не найдены на этой странице. Завершаем парсинг.")
                    break  # Если на странице нет акций, выходим

                for promo in tqdm(promo_blocks, desc=f"Обрабатываем акции Mechta:"):
                    title_tag = promo.find("p", class_="text-color3 text-ts3")  # Название акции
                    date_tag = promo.find("p", class_="text-color1 text-ts1 text-bold")  # Даты
                    description_tag = promo.find("p", class_="text-color1 text-ts3 col")  # Описание
                    
                    # Находим ближайшую ссылку
                    link_tag = promo.find_next("a", href=True)  # Ищем следующий тег <a>
                    link = link_tag['href'] if link_tag else "Нет ссылки"

                    if title_tag:
                        all_promotions.append([
                            title_tag.get_text(strip=True),
                            description_tag.get_text(strip=True) if description_tag else "Нет описания",
                            date_tag.get_text(strip=True) if date_tag else "Не указано",
                            link
                        ])

                page += 1  # Увеличиваем номер страницы

            driver.quit()
            return all_promotions

        except Exception as e:
            print(f"❌ Ошибка с Selenium: {e}")
            return []



    def parse_technodom(self):
        """Парсит акции с Technodom.kz (через Selenium)."""
        print("[🔍 Парсинг Technodom]")

        try:
            # Запуск Selenium
            driver = webdriver.Chrome(service=self.service, options=self.chrome_options)
            driver.get(self.TECHNODOM_URL)

            # Ждём загрузку хотя бы одного блока с акцией
            driver.implicitly_wait(5)

            promotions = []
            
            # Найти все промо-блоки
            promo_blocks = driver.find_elements(By.CLASS_NAME, "promo__info")
            
            for promo in  tqdm(promo_blocks, desc="Обрабатываем акции Technodom:"):
                # Извлекаем заголовок акции
                heading = promo.find_element(By.CLASS_NAME, "promo__heading").text
                # Извлекаем описание акции
                text = promo.find_element(By.CLASS_NAME, "promo__text").text
                # Извлекаем ссылку
                link = promo.find_element(By.CLASS_NAME, "promo__heading-link").get_attribute("href")

                promotions.append([heading, text, link])

            driver.quit()
            return promotions

        except Exception as e:
            print(f"❌ Ошибка при парсинге Technodom: {e}")
            return []



    def save_to_excel(self, sulpak_data, mechta_data, technodom_data):
        """Сохраняет данные в Excel-файл с разными вкладками."""
        workbook = openpyxl.Workbook()

        # 🔴 Лист Sulpak
        sheet_sulpak = workbook.active
        sheet_sulpak.title = "Sulpak"
        sheet_sulpak.append(["Название акции", "Описание", "Период акции"])
        for row in sulpak_data:
            sheet_sulpak.append(row)

        # 🔵 Лист Mechta (теперь со ссылками!)
        sheet_mechta = workbook.create_sheet(title="Mechta")
        sheet_mechta.append(["Период акции", "Название акции", "Описание", "Ссылка"])  # Добавляем колонку для ссылки

        for row in mechta_data:
            period, title, description, link = row  # Извлекаем данные
            
            if link.startswith("http"):
                sheet_mechta.append([period, title, description, link])  # Добавляем данные
                last_row = sheet_mechta.max_row
                link_cell = sheet_mechta.cell(row=last_row, column=4)  # Колонка со ссылкой
                link_cell.hyperlink = link  # Делаем ссылку кликабельной
                link_cell.font = Font(color="0000FF", underline="single")  # Стилизуем как ссылку
            else:
                sheet_mechta.append([period, title, description, "Нет ссылки"])  # Если ссылки нет

        # 🟢 Лист Technodom (с кликабельными ссылками)
        sheet_technodom = workbook.create_sheet(title="Technodom")
        sheet_technodom.append(["Название акции", "Описание", "Ссылка"])

        for row in technodom_data:
            title, description, link = row

            # Проверяем, является ли ссылка корректной
            if link.startswith("http"):
                cell = sheet_technodom.append([title, description, link])  # Добавляем ссылку в ячейку
                last_row = sheet_technodom.max_row
                link_cell = sheet_technodom.cell(row=last_row, column=3)
                link_cell.hyperlink = link  # Делаем ячейку гиперссылкой
                link_cell.font = Font(color="0000FF", underline="single")  # Стилизуем как ссылку
            else:
                sheet_technodom.append([title, description, "Нет ссылки"])  # Если ссылки нет


        try:
            # 📌 **Сохраняем файл в память, а не на диск**
            # output = BytesIO()
            # workbook.save(output)
            # output.seek(0)  # Перемещаем указатель в начало файла (важно!)
            # return output  # Возвращаем объект файла в памяти
            workbook.save(self.file_path)
            print(f"✅ Файл сохранён: {self.file_path}")
        except Exception as e:
            print(f"❌ Ошибка при сохранении файла: {e}")


    def run(self):
        """Запускает парсинг и возвращает путь к Excel-файлу."""
        sulpak_promotions = self.parse_sulpak()
        mechta_promotions = self.parse_mechta()
        technodom_promotions = self.parse_technodom()
        return self.save_to_excel(sulpak_promotions, mechta_promotions, technodom_promotions)
    

# ✅ **Запускаем парсер, если файл запущен напрямую**
if __name__ == "__main__":
    parser = PromoParser()
    excel_file = parser.run()
   


    # **************************************************************** функция для использования кода в качестве класса. вставить в main function
    # from Parcer import PromoParser  # Импортируем наш парсер

    # def send_promotions(update: Update, context: CallbackContext):
    #     """Отправляет пользователю файл с акциями в виде Excel-документа (без сохранения на диск)."""
    #     parser = PromoParser()
    #     excel_file = parser.run()  # Получаем файл в памяти (BytesIO)
    #     update.message.reply_document(document=excel_file, filename="promotions.xlsx")